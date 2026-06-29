"""
Utilities for reading and writing Excel (.xlsx) workbooks.
"""

from __future__ import annotations

import json
import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

_SHEET_NAME_SAFE = re.compile(r"[\[\]\:\*\?\/\\]+")
_MAX_SHEET_NAME_LEN = 31
_MAX_ROWS_PER_SHEET = 10_000
_MAX_COLS_PER_SHEET = 256


def _sanitize_sheet_name(name: str, fallback: str) -> str:
    cleaned = _SHEET_NAME_SAFE.sub("_", (name or "").strip())[:_MAX_SHEET_NAME_LEN]
    return cleaned or fallback


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def extract_xlsx_text_from_bytes(raw: bytes) -> str:
    """
    Convert an .xlsx workbook into sheet-aware plain text for RAG indexing.

    Each sheet is rendered as:
      === Sheet: <name> ===
      col1 | col2 | ...
    """
    workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    parts: list[str] = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [_cell_to_text(cell) for cell in row]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))

    workbook.close()
    return "\n\n".join(parts).strip()


def build_xlsx_bytes_from_sheets(sheets: list[dict[str, Any]]) -> bytes:
    """
    Build an .xlsx workbook from structured sheet definitions.

    Each sheet dict supports:
      - name: str
      - headers: list[str] (optional)
      - rows: list[list[Any]] (optional)
    """
    if not sheets:
        raise ValueError("At least one sheet is required")

    workbook = Workbook()
    workbook.remove(workbook.active)

    for index, sheet_def in enumerate(sheets):
        if not isinstance(sheet_def, dict):
            raise ValueError(f"sheets[{index}] must be an object")

        sheet_name = _sanitize_sheet_name(
            str(sheet_def.get("name") or ""),
            fallback=f"Sheet{index + 1}",
        )
        worksheet = workbook.create_sheet(title=sheet_name)

        headers = sheet_def.get("headers") or []
        if headers and not isinstance(headers, list):
            raise ValueError(f"sheets[{index}].headers must be a list")
        if headers:
            worksheet.append([_cell_to_text(h) for h in headers[:_MAX_COLS_PER_SHEET]])

        rows = sheet_def.get("rows") or []
        if rows and not isinstance(rows, list):
            raise ValueError(f"sheets[{index}].rows must be a list")

        row_count = 0
        for row in rows:
            if row_count >= _MAX_ROWS_PER_SHEET:
                break
            if not isinstance(row, list):
                raise ValueError(f"sheets[{index}].rows entries must be lists")
            worksheet.append(
                [_cell_to_text(cell) for cell in row[:_MAX_COLS_PER_SHEET]]
            )
            row_count += 1

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_sheets_json(sheets_json: str) -> list[dict[str, Any]]:
    """Parse a JSON string into sheet definitions for workbook creation."""
    try:
        parsed = json.loads(sheets_json or "[]")
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid sheets JSON: {exc}") from exc

    if not isinstance(parsed, list):
        raise ValueError("sheets JSON must be a list of sheet objects")
    if not parsed:
        raise ValueError("sheets JSON must contain at least one sheet")
    return parsed
